# coding:utf-8
from __future__ import division
from pymongo import MongoClient
import jieba
import jieba.posseg as pseg
from openpyxl import load_workbook
from openpyxl import Workbook
import re
jieba.load_userdict('E:/博二/奇瑞项目/项目代码/userdict.txt')

#####################################  分割线  #####################################

#载入产品特征词以及观点词
def load_feature_senword_dataSet(filename):
    wb=load_workbook(filename)
    ws1 = wb['featureSet']
    ws2 = wb['senwordSet']
    featureSet=[]
    senwordSet=[]
    #senWeight=[]
    degwordSet=[]
    #degWeight=[]
    notwordSet=[]
    
    for i in range(1,ws1.max_row+1):
        feature=ws1.cell(row=i,column=1).value
        featureSet.append(feature)

    for i in range(1,ws2.max_row+1):
        senword=ws2.cell(row=i,column=1).value
        #polarity=ws2.cell(row=i,column=2).value
        senwordSet.append(senword)
        #senWeight.append(polarity)
        
    for i in range(1,203):
        degword=ws2.cell(row=i,column=3).value
        #weight=ws2.cell(row=i,column=4).value
        degwordSet.append(degword)
        #degWeight.append(weight)

    for i in range(1,21):
        notword=ws2.cell(row=i,column=5).value
        notwordSet.append(notword)
    return featureSet,senwordSet,degwordSet,notwordSet

#将长句切分为短句
def clauseSegmentation(comment):
    reg1=re.compile(u'。+')
    reg2=re.compile(u'[？?]+')
    reg3=re.compile(u'[！!]+')
    reg4=re.compile(u'[;；]+')
    reg5=re.compile(u'\d[、,，：:]+')
    reg6=re.compile(u'…+')
    reg7=re.compile(u'（\d）')
    reg9=re.compile(u'，+')
    reg10=re.compile(u'\t+')
    comment=reg1.sub(u'。【Instead】',comment)
    comment=reg2.sub(u'？【Instead】',comment)
    comment=reg3.sub(u'！【Instead】',comment)
    comment=reg4.sub(u'；【Instead】',comment)
    comment=reg5.sub(u'【Instead】',comment)
    comment=reg6.sub(u'……【Instead】',comment)
    comment=reg7.sub(u'',comment)
    comment=reg9.sub('，',comment)
    comment=reg10.sub('【Instead】',comment)
    clauseSet=re.split(u'【Instead】',comment)
    clauseList=[clause for clause in clauseSet if len(clause)>5]
    return clauseList
#找到某个特征词对应的所有位置
def myfind(x,y):
    return [ a for a in range(len(y)) if y[a] == x]

def deepAnalysis(f_index,index2,senword,degwordSet,notwordSet,clauseVecSet):
    opinion=''
    degword=''
    notword=''
    degIndex=0
    notIndex=0
    for x in range(f_index+1,index2):
        if clauseVecSet[x][0] in degwordSet:
            degword=clauseVecSet[x][0]
            degIndex=x
        if clauseVecSet[x][0] in notwordSet:
            notword=clauseVecSet[x][0]
            notIndex=x
    if degIndex>notIndex:
        opinion=notword+degword+senword
    if notIndex>=degIndex:
        opinion=degword+notword+senword
    return opinion

#从左边找，停止条件：1.遇到逗号；2.抵达句子开始；3.遇到其它特征词;4.不需要程度词和否定词
def forwardSearch(f_index,featureSet,senwordSet,clauseVecSet):
    opinion=''
    for k in range(1,6):
        
        index1=f_index-k
        if index1>=0:
            if clauseVecSet[index1][1] == 'x':break
            if clauseVecSet[index1][0] in featureSet:break
            if (clauseVecSet[index1][0] in senwordSet) or (clauseVecSet[index1][1] =='a'):
                opinion=clauseVecSet[index1][0]
                break
    return opinion

#从右边找，停止条件：1.遇到逗号；2.抵达句子末端；3.遇到其它特征词；4.需要程度词和否定词
def backwardSearch(f_index,featureSet,senwordSet,degwordSet,notwordSet,clauseVecSet):
    opinion=''
    for k in range(1,6):
        index2=f_index+k
        if index2<len(clauseVecSet):
            if clauseVecSet[index2][1] == 'x':break
            if clauseVecSet[index2][0] in featureSet:break
            if (clauseVecSet[index2][0] in senwordSet) or (clauseVecSet[index2][1] =='a'):
                senword=clauseVecSet[index2][0]
                #定位程度词与否定词
                opinion=deepAnalysis(f_index,index2,senword,degwordSet,notwordSet,clauseVecSet)
                break
    return opinion

def splitDataSet(dataSet, axis, value):#按照给定特征划分数据集
    retDataSet = []#用来存储划分结果的列表
    for featVec in dataSet:#对于每行数据
        if featVec[axis] == value:#如果选择的特征等于给定值
            reducedFeatVec = featVec[:axis]
            reducedFeatVec.extend(featVec[axis+1:])
            retDataSet.append(reducedFeatVec)#将这行数据存入划分结果的列表中（不包含给定特征）
    return retDataSet#返回划分结果数据集

#将数据格式化为字典
def formatDataSet(opinionSet):
    formatOpinionSet={}
    feaUniqueVals=set([exa[0] for exa in opinionSet])
    for value1 in feaUniqueVals:
        subDataSet=splitDataSet(opinionSet,0,value1)
        senUniqueVals=set([exa[0] for exa in subDataSet])
        subDic={}
        for value2 in senUniqueVals:
            commentSet=[c[0] for c in splitDataSet(subDataSet,0,value2)]
            subDic[value2]=commentSet
        formatOpinionSet[value1]=subDic
    return formatOpinionSet

def opinionMining(comment,featureSet,senwordSet,degwordSet,notwordSet):
    #分句
    clauseList=clauseSegmentation(comment)
    opinionSet=[]
    #对于每一条子句，分别提取特征词与观点词
    for i in range(len(clauseList)):
        clause=clauseList[i]
        clauseVecSet=[(w.word,w.flag) for w in pseg.cut(clause)]
        clauseWordSet=[w[0] for w in clauseVecSet]
        #print(clauseVecSet)
        for j in range(len(featureSet)):
            #定位特征词
            feature=featureSet[j]
            f_index=myfind(feature,clauseWordSet)
            if len(f_index)>0:index=f_index[0]
            else:continue
            #定位观点词
            opinion=forwardSearch(index,featureSet,senwordSet,clauseVecSet)
            if len(opinion)==0:
                opinion=backwardSearch(index,featureSet,senwordSet,degwordSet,notwordSet,clauseVecSet)
            if len(opinion)==0:continue
            result=[feature,opinion,clause]
            opinionSet.append(result)
    formatOpinionSet=formatDataSet(opinionSet)
    return formatOpinionSet

#####################################  分割线  #####################################

def test():
    featureSet,senwordSet,degwordSet,notwordSet=load_feature_senword_dataSet('E:/博二/奇瑞项目/项目代码/产品特征-情感词_数据集.xlsx')
    wb=Workbook()
    ws=wb.active

    client = MongoClient('47.99.116.136', 30000)
    collection = client.car.comment
    print('一共包含数据%s条'%collection.find().count())
    #获取所有id
    _idSet=collection.find({},{"_id":1}).batch_size(1000)
    #读取每一个id对应的原始数据
    n=0
    m=1
    for _id in _idSet:
        if n%100==0:print(n)
        if n==20000:break
        mongoData=list(collection.find({"_id":_id['_id']},{"data":1,"url":1,"status":1}))[0]
        #分别对每条原始数据进行处理，得到中间数据
        dataSource=mongoData["url"]
        try:
            place=mongoData["data"]["购买地点"]
        except:
            place=''
        car=mongoData["data"]["购买车型"]
        commentSet=mongoData["data"]["评价"]
        whole_featureSet=list(commentSet.keys())
        for i in range(len(whole_featureSet)):
            whole_feature=whole_featureSet[i]
            comment=commentSet[whole_feature]
            opinionSet=opinionMining(comment,featureSet,senwordSet,degwordSet,notwordSet)
            for j in range(len(opinionSet)):
                ws.cell(row=m,column=1).value=m
                ws.cell(row=m,column=2).value=dataSource
                ws.cell(row=m,column=3).value=place
                ws.cell(row=m,column=4).value=car
                ws.cell(row=m,column=5).value=whole_feature
                ws.cell(row=m,column=6).value=opinionSet[j][0]
                ws.cell(row=m,column=7).value=opinionSet[j][1]
                ws.cell(row=m,column=8).value=opinionSet[j][2]
                m+=1
        n+=1
        if n%2000==0:
            print('原始数据数量%s'%n)
            print('生成数据数量%s'%(m-1))
    wb.save('result.xlsx')

#featureSet,senwordSet,degwordSet,notwordSet=load_feature_senword_dataSet('E:/博二/奇瑞项目/项目代码/产品特征-情感词_数据集.xlsx')
#resultSet=opinionMining('内饰太优秀了,外观大气，性价比高！整体感觉良好，再说一遍外观好',featureSet,senwordSet,degwordSet,notwordSet)
#print(resultSet)

#clause='外观真的太差了。'
#clauseVecSet=[(w.word,w.flag) for w in pseg.cut(clause)]
#print(clauseVecSet)
#opinion=deepAnalysis(0,2,'太差',degwordSet,notwordSet,clauseVecSet)
#print(opinion)
