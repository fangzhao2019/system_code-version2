# coding:utf-8
from pymongo import MongoClient
from openpyxl import Workbook

wb=Workbook()
ws=wb.active

client=MongoClient('47.92.211.251',30000)
collection=client.car.car_config

print('一共包含数据%s条'%collection.find().count())

_idSet=collection.find({},{"_id":1}).batch_size(1000)

m=0
carConfigSet=[]
totalKey=[]
for _id in _idSet:
    m+=1
    if m%10==0:print(m)
    mongoData=list(collection.find({"_id":_id['_id']},{"carTypes":1}))[0]['carTypes']
    if len(mongoData)==0:continue
    #遍历每款车下的小车型
    for coarseConfig in mongoData:
        carConfig={}
        carConfig['car']=coarseConfig['配置名']
        carConfig['price']=coarseConfig['官方价']
        for key1 in coarseConfig.keys():
            carFineConfig=coarseConfig[key1]
            if type(carFineConfig)==type({}):       
                for key2 in carFineConfig.keys():
                    carConfig[key2]=carFineConfig[key2]
                    if not key2 in totalKey:
                        totalKey.append(key2)
        carConfigSet.append(carConfig)
print('一共包含配置%d个'%len(totalKey))
ws.cell(row=1,column=1).value='车型'
ws.cell(row=1,column=2).value='价格'
for k in range(len(totalKey)):
    ws.cell(row=1,column=k+3).value=totalKey[k]
    
i=1   
#遍历每款车下的小车型
for carConfig in carConfigSet:
    i+=1
    if i%100==0:print(i)
    ws.cell(row=i,column=1).value=carConfig['car']
    ws.cell(row=i,column=2).value=carConfig['price']
    j=1
    for key in totalKey:
        if key in carConfig.keys():
            ws.cell(row=i,column=j+2).value=carConfig[key]
        else:
            ws.cell(row=i,column=j+2).value='-'
        j+=1

wb.save('carConfig.xlsx')
