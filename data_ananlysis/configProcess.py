# coding:utf-8
from openpyxl import load_workbook
import numpy as np
from openpyxl import Workbook

#载入配置信息
wb=load_workbook('carConfig.xlsx')
ws=wb.active

new_wb=Workbook()
new_ws=new_wb.active

m=1
for j in range(3,ws.max_column+1):
	configName=ws.cell(row=1,column=j).value
	configSet={}
	for i in range(2,ws.max_row+1):
		config=ws.cell(row=i,column=j).value
		if not config in configSet.keys():
			configSet[config]=0
		configSet[config]+=1
	configFormat=sorted(configSet.items(),key = lambda x:x[1],reverse = True)
	if len(configFormat)<2:continue
	if '-' in configSet.keys():
		radio=configSet['-']/float(ws.max_row-1)
	else:
		radio=0.0
	if radio>0.8:continue

	new_ws.cell(row=m,column=1).value=configName
	new_ws.cell(row=m,column=2).value=radio
	new_ws.cell(row=m,column=3).value=str(configSet.keys())
	m+=1
new_wb.save('configInfo.xlsx')
	#print('%s %.6f %s'%(configName,radio,str(configFormat.keys()))